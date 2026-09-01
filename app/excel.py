"""엑셀(.xlsx) 업로드·다운로드 처리."""

from __future__ import annotations

import io
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import (
    ASSET_CATEGORIES,
    ASSET_STATUSES,
    EMPLOYEE_STATUSES,
    Asset,
    Assignment,
    Employee,
)

HEADER_FILL = PatternFill("solid", fgColor="1F3B63")
HEADER_FONT = Font(color="FFFFFF", bold=True)

# 업로드 시 한글 헤더를 모델 필드로 연결한다.
ASSET_COLUMNS: list[tuple[str, str, int]] = [
    ("asset_no", "자산번호", 16),
    ("name", "자산명", 26),
    ("category", "분류", 14),
    ("status", "상태", 10),
    ("manufacturer", "제조사", 14),
    ("model_name", "모델명", 20),
    ("serial_no", "시리얼번호", 20),
    ("spec", "사양", 30),
    ("location", "보관위치", 16),
    ("supplier", "공급업체", 16),
    ("purchase_date", "도입일", 12),
    ("purchase_price", "취득가액", 14),
    ("warranty_until", "보증만료일", 12),
    ("license_key", "라이선스키", 24),
    ("holder_emp_no", "사용자 사번", 14),
    ("holder_name", "사용자명", 12),
    ("note", "비고", 30),
]

# 업로드 파일에서 필수인 열
REQUIRED_UPLOAD_HEADERS = ("자산번호", "자산명")

_CATEGORY_BY_LABEL = {label: code for code, label in ASSET_CATEGORIES.items()}
_STATUS_BY_LABEL = {label: code for code, label in ASSET_STATUSES.items()}


# --- 공통 유틸 ----------------------------------------------------------------

def _style_header(ws, headers: list[str], widths: list[int]) -> None:
    ws.append(headers)
    for idx, (header, width) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=1, column=idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A2"


def _to_bytes(wb: Workbook) -> bytes:
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _cell_text(value: Any) -> str | None:
    """셀 값을 문자열로 정리. 빈 값은 None."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float) and value.is_integer():
        text = str(int(value))
    else:
        text = str(value)
    text = text.strip()
    return text or None


def _cell_date(value: Any, label: str) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y%m%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"{label} 형식이 올바르지 않습니다: '{text}' (예: 2026-01-31)")


def _cell_money(value: Any, label: str) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        amount = float(value)
    else:
        text = str(value).replace(",", "").replace("원", "").strip()
        if not text:
            return None
        try:
            amount = float(text)
        except ValueError as exc:
            raise ValueError(f"{label}은(는) 숫자로 입력해 주세요: '{value}'") from exc
    if amount < 0:
        raise ValueError(f"{label}은(는) 0 이상이어야 합니다.")
    return amount


def _cell_code(value: Any, labels: dict[str, str], by_label: dict[str, str], field_label: str, default: str) -> str:
    """'노트북' 또는 'NOTEBOOK' 어느 쪽으로 적어도 코드로 변환한다."""
    text = _cell_text(value)
    if text is None:
        return default
    if text in labels:
        return text
    if text.upper() in labels:
        return text.upper()
    if text in by_label:
        return by_label[text]
    allowed = ", ".join(labels.values())
    raise ValueError(f"{field_label} 값이 올바르지 않습니다: '{text}' (가능한 값: {allowed})")


# --- 다운로드 -----------------------------------------------------------------

def export_assets(assets: list[Asset]) -> bytes:
    """자산 목록을 엑셀 파일 바이트로."""
    wb = Workbook()
    ws = wb.active
    ws.title = "자산목록"
    _style_header(ws, [label for _, label, _ in ASSET_COLUMNS], [w for _, _, w in ASSET_COLUMNS])

    for asset in assets:
        ws.append(
            [
                asset.asset_no,
                asset.name,
                asset.category_label,
                asset.status_label,
                asset.manufacturer,
                asset.model_name,
                asset.serial_no,
                asset.spec,
                asset.location,
                asset.supplier,
                asset.purchase_date,
                float(asset.purchase_price) if asset.purchase_price is not None else None,
                asset.warranty_until,
                asset.license_key,
                asset.holder.emp_no if asset.holder else None,
                asset.holder.name if asset.holder else None,
                asset.note,
            ]
        )

    for row in ws.iter_rows(min_row=2):
        row[10].number_format = "yyyy-mm-dd"  # 도입일
        row[11].number_format = "#,##0"       # 취득가액
        row[12].number_format = "yyyy-mm-dd"  # 보증만료일
    return _to_bytes(wb)


def export_asset_template() -> bytes:
    """업로드용 빈 양식 (예시 1행 포함)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "자산등록양식"
    _style_header(ws, [label for _, label, _ in ASSET_COLUMNS], [w for _, _, w in ASSET_COLUMNS])
    ws.append(
        [
            "IT-2026-0001", "개발팀 노트북", "노트북", "재고", "LG전자", "그램 16",
            "SN-EXAMPLE-001", "i7 / 32GB / 1TB", "본사 3층 창고", "테크상사",
            "2026-01-15", 2150000, "2029-01-14", "", "", "", "예시 행입니다. 지우고 사용하세요.",
        ]
    )

    guide = wb.create_sheet("작성안내")
    guide.column_dimensions["A"].width = 18
    guide.column_dimensions["B"].width = 90
    guide.append(["항목", "설명"])
    for cell in guide[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    rows = [
        ("자산번호", "필수. 시스템에서 자산을 구분하는 고유 번호입니다. 이미 있는 번호면 해당 자산 정보가 갱신됩니다."),
        ("자산명", "필수. 예: 개발팀 노트북"),
        ("분류", f"다음 중 하나: {', '.join(ASSET_CATEGORIES.values())} (비우면 '기타')"),
        ("상태", f"다음 중 하나: {', '.join(ASSET_STATUSES.values())} (비우면 '재고')"),
        ("도입일 / 보증만료일", "YYYY-MM-DD 형식으로 입력하세요. 예: 2026-01-31"),
        ("취득가액", "숫자만 입력하세요. 쉼표는 넣어도 됩니다."),
        ("사용자 사번", "이미 등록된 직원의 사번을 넣으면 해당 직원에게 지급 처리되고 이력이 남습니다. 비우면 미지급 상태."),
        ("사용자명", "참고용입니다. 지급 대상은 '사용자 사번'으로 판단합니다."),
    ]
    for row in rows:
        guide.append(row)
        guide.cell(row=guide.max_row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    return _to_bytes(wb)


def export_employees(employees: list[Employee]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "직원목록"
    headers = ["사번", "이름", "부서", "직급", "이메일", "연락처", "재직상태", "보유자산수", "비고"]
    _style_header(ws, headers, [14, 12, 16, 12, 26, 16, 10, 12, 30])
    for emp in employees:
        ws.append(
            [
                emp.emp_no,
                emp.name,
                emp.department,
                emp.position,
                emp.email,
                emp.phone,
                emp.status_label,
                len(emp.assets),
                emp.note,
            ]
        )
    return _to_bytes(wb)


def export_assignments(assignments: list[Assignment]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "지급반납이력"
    headers = [
        "자산번호", "자산명", "분류", "사번", "이름", "부서",
        "지급일", "반납일", "보유일수", "상태", "지급메모", "반납메모", "처리자",
    ]
    _style_header(ws, headers, [16, 24, 12, 14, 12, 16, 12, 12, 10, 10, 24, 24, 12])
    for item in assignments:
        ws.append(
            [
                item.asset.asset_no,
                item.asset.name,
                item.asset.category_label,
                item.employee.emp_no,
                item.employee.name,
                item.employee.department,
                item.assigned_at,
                item.returned_at,
                item.days_held,
                "지급중" if item.is_open else "반납완료",
                item.assigned_note,
                item.return_note,
                item.created_by,
            ]
        )
    for row in ws.iter_rows(min_row=2):
        row[6].number_format = "yyyy-mm-dd"
        row[7].number_format = "yyyy-mm-dd"
    return _to_bytes(wb)


# --- 업로드 -------------------------------------------------------------------

@dataclass
class ImportResult:
    """엑셀 업로드 결과 요약."""

    created: int = 0
    updated: int = 0
    skipped: int = 0
    errors: list[str] = field(default_factory=list)

    @property
    def total_processed(self) -> int:
        return self.created + self.updated

    @property
    def has_errors(self) -> bool:
        return bool(self.errors)

    def summary(self) -> str:
        return (
            f"신규 {self.created}건, 수정 {self.updated}건, 실패 {self.skipped}건을 처리했습니다."
        )


MAX_REPORTED_ERRORS = 30


def import_assets(db: Session, content: bytes, *, actor: str | None = None) -> ImportResult:
    """엑셀 파일을 읽어 자산을 등록/갱신한다.

    자산번호가 이미 있으면 갱신, 없으면 신규 등록한다.
    한 행이 잘못돼도 나머지 행은 계속 처리하고, 오류는 결과에 모아 돌려준다.
    """
    result = ImportResult()

    try:
        wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    except Exception as exc:  # openpyxl 은 다양한 예외를 낸다
        result.errors.append(f"엑셀 파일을 열 수 없습니다: {exc}")
        result.skipped = 1
        return result

    ws = wb.worksheets[0]
    rows = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows)
    except StopIteration:
        result.errors.append("빈 파일입니다. 데이터가 있는 엑셀을 올려 주세요.")
        return result

    headers = [(_cell_text(c) or "") for c in header_row]
    index = {header: pos for pos, header in enumerate(headers) if header}
    missing = [h for h in REQUIRED_UPLOAD_HEADERS if h not in index]
    if missing:
        result.errors.append(
            f"필수 열이 없습니다: {', '.join(missing)}. '양식 다운로드'로 받은 파일을 사용해 주세요."
        )
        return result

    def value_of(row: tuple, header: str):
        pos = index.get(header)
        if pos is None or pos >= len(row):
            return None
        return row[pos]

    # 사번으로 직원을 찾기 위한 캐시
    employees = {
        emp.emp_no: emp for emp in db.scalars(select(Employee)).all()
    }
    seen_asset_nos: set[str] = set()

    for row_no, row in enumerate(rows, start=2):
        if row is None or all(cell is None or str(cell).strip() == "" for cell in row):
            continue  # 빈 행은 건너뛴다

        try:
            asset_no = _cell_text(value_of(row, "자산번호"))
            if not asset_no:
                raise ValueError("자산번호가 비어 있습니다.")
            if asset_no in seen_asset_nos:
                raise ValueError(f"같은 파일 안에 자산번호 '{asset_no}'가 중복으로 있습니다.")
            seen_asset_nos.add(asset_no)

            name = _cell_text(value_of(row, "자산명"))
            if not name:
                raise ValueError("자산명이 비어 있습니다.")

            asset = db.scalar(select(Asset).where(Asset.asset_no == asset_no))
            is_new = asset is None
            if is_new:
                asset = Asset(asset_no=asset_no)
                db.add(asset)

            asset.name = name
            asset.category = _cell_code(
                value_of(row, "분류"), ASSET_CATEGORIES, _CATEGORY_BY_LABEL, "분류", "ETC"
            )
            status = _cell_code(
                value_of(row, "상태"), ASSET_STATUSES, _STATUS_BY_LABEL, "상태", "IN_STOCK"
            )
            asset.manufacturer = _cell_text(value_of(row, "제조사"))
            asset.model_name = _cell_text(value_of(row, "모델명"))
            asset.serial_no = _cell_text(value_of(row, "시리얼번호"))
            asset.spec = _cell_text(value_of(row, "사양"))
            asset.location = _cell_text(value_of(row, "보관위치"))
            asset.supplier = _cell_text(value_of(row, "공급업체"))
            asset.purchase_date = _cell_date(value_of(row, "도입일"), "도입일")
            asset.purchase_price = _cell_money(value_of(row, "취득가액"), "취득가액")
            asset.warranty_until = _cell_date(value_of(row, "보증만료일"), "보증만료일")
            asset.license_key = _cell_text(value_of(row, "라이선스키"))
            asset.note = _cell_text(value_of(row, "비고"))

            holder_emp_no = _cell_text(value_of(row, "사용자 사번"))
            db.flush()  # 신규 자산의 id 를 확보한다

            if holder_emp_no:
                employee = employees.get(holder_emp_no)
                if employee is None:
                    raise ValueError(
                        f"사번 '{holder_emp_no}' 인 직원이 없습니다. 직원을 먼저 등록해 주세요."
                    )
                _sync_holder(db, asset, employee, actor)
            else:
                asset.status = status
                if not is_new and asset.holder_id is not None and status != "IN_USE":
                    # 사용자 열을 비워서 올리면 반납 처리로 본다
                    _close_open_assignment(db, asset, actor)

            result.created += int(is_new)
            result.updated += int(not is_new)
        except Exception as exc:  # 한 행의 오류가 전체를 막지 않게 한다
            db.rollback()
            result.skipped += 1
            if len(result.errors) < MAX_REPORTED_ERRORS:
                result.errors.append(f"{row_no}행: {exc}")
            continue
        else:
            db.commit()

    wb.close()
    if result.skipped > MAX_REPORTED_ERRORS:
        result.errors.append(f"... 그 외 {result.skipped - MAX_REPORTED_ERRORS}건의 오류가 더 있습니다.")
    return result


def _close_open_assignment(db: Session, asset: Asset, actor: str | None) -> None:
    """열려 있는 지급 이력을 반납 처리한다 (자산 상태는 호출한 쪽에서 정한다)."""
    from app.services import open_assignment

    current = open_assignment(db, asset.id)
    if current is not None:
        current.returned_at = date.today()
        current.return_note = "엑셀 업로드로 반납 처리"
        current.returned_by = actor
    asset.holder_id = None


def _sync_holder(db: Session, asset: Asset, employee: Employee, actor: str | None) -> None:
    """엑셀의 '사용자 사번'에 맞춰 지급 상태를 맞춘다."""
    from app.services import open_assignment

    current = open_assignment(db, asset.id)
    if current is not None and current.employee_id == employee.id:
        asset.status = "IN_USE"
        asset.holder_id = employee.id
        return

    if current is not None:
        # 사용자가 바뀌었으면 기존 건을 반납 처리하고 새로 지급한다
        current.returned_at = date.today()
        current.return_note = "엑셀 업로드로 사용자 변경"
        current.returned_by = actor

    # 지급일은 도입일로 잡되, 도입일이 미래이거나 비어 있으면 오늘로 둔다.
    today = date.today()
    assigned_at = min(asset.purchase_date, today) if asset.purchase_date else today

    db.add(
        Assignment(
            asset_id=asset.id,
            employee_id=employee.id,
            assigned_at=assigned_at,
            assigned_note="엑셀 업로드로 지급 처리",
            created_by=actor,
        )
    )
    asset.status = "IN_USE"
    asset.holder_id = employee.id
