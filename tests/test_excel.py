"""엑셀 업로드 / 다운로드 테스트."""

import io
from datetime import date

from openpyxl import Workbook, load_workbook
from sqlalchemy import select

from app.excel import ASSET_COLUMNS, export_assets, import_assets
from app.models import Asset, Assignment, Employee

HEADERS = [label for _, label, _ in ASSET_COLUMNS]


def _upload_file(rows: list[list], headers: list[str] | None = None) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.append(headers or HEADERS)
    for row in rows:
        ws.append(row)
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _row(asset_no="IT-2026-0001", name="업로드 노트북", **cells) -> list:
    values = {
        "자산번호": asset_no,
        "자산명": name,
        "분류": "노트북",
        "상태": "재고",
        "제조사": "LG전자",
        "모델명": "그램 16",
        "도입일": "2026-01-15",
        "취득가액": 2150000,
        "보증만료일": "2029-01-14",
    }
    values.update(cells)
    return [values.get(header) for header in HEADERS]


def test_엑셀_업로드로_신규_자산이_등록된다(db):
    result = import_assets(db, _upload_file([_row()]))

    assert result.created == 1 and result.skipped == 0
    asset = db.scalar(select(Asset).where(Asset.asset_no == "IT-2026-0001"))
    assert asset.name == "업로드 노트북"
    assert asset.category == "NOTEBOOK"
    assert asset.purchase_date == date(2026, 1, 15)
    assert float(asset.purchase_price) == 2150000


def test_같은_자산번호는_새로_만들지_않고_갱신한다(db, asset):
    result = import_assets(db, _upload_file([_row(asset_no=asset.asset_no, name="이름 변경됨")]))

    assert result.created == 0 and result.updated == 1
    db.expire_all()
    assert db.get(Asset, asset.id).name == "이름 변경됨"


def test_영문_코드로_적어도_인식한다(db):
    result = import_assets(db, _upload_file([_row(**{"분류": "SOFTWARE", "상태": "IN_USE"})]))
    assert result.created == 1
    assert db.scalar(select(Asset)).category == "SOFTWARE"


def test_잘못된_행은_건너뛰고_나머지는_반영된다(db):
    content = _upload_file(
        [
            _row(asset_no="IT-OK-1"),
            _row(asset_no="", name="자산번호 없음"),
            _row(asset_no="IT-BAD", **{"분류": "없는분류"}),
            _row(asset_no="IT-OK-2"),
        ]
    )
    result = import_assets(db, content)

    assert result.created == 2
    assert result.skipped == 2
    assert len(result.errors) == 2
    assert db.scalar(select(Asset).where(Asset.asset_no == "IT-BAD")) is None


def test_한_파일에_같은_자산번호가_두_번_나오면_오류(db):
    result = import_assets(db, _upload_file([_row(asset_no="IT-DUP"), _row(asset_no="IT-DUP")]))
    assert result.created == 1
    assert result.skipped == 1
    assert "중복" in result.errors[0]


def test_사용자_사번을_채우면_지급_이력이_생긴다(db, employee):
    result = import_assets(
        db, _upload_file([_row(**{"사용자 사번": employee.emp_no})]), actor="admin"
    )

    assert result.created == 1
    asset = db.scalar(select(Asset).where(Asset.asset_no == "IT-2026-0001"))
    assert asset.status == "IN_USE"
    assert asset.holder_id == employee.id

    assignment = db.scalar(select(Assignment).where(Assignment.asset_id == asset.id))
    assert assignment is not None and assignment.is_open


def test_없는_사번을_적으면_해당_행만_실패한다(db):
    result = import_assets(db, _upload_file([_row(**{"사용자 사번": "없는사번"})]))
    assert result.skipped == 1
    assert "직원이 없습니다" in result.errors[0]


def test_사용자_사번을_바꾸면_기존_지급건이_반납되고_새로_지급된다(db, employee):
    other = Employee(emp_no="E002", name="김철수", department="영업팀")
    db.add(other)
    db.commit()

    import_assets(db, _upload_file([_row(**{"사용자 사번": employee.emp_no})]))
    import_assets(db, _upload_file([_row(**{"사용자 사번": other.emp_no})]))

    asset = db.scalar(select(Asset).where(Asset.asset_no == "IT-2026-0001"))
    assert asset.holder_id == other.id

    history = db.scalars(select(Assignment).where(Assignment.asset_id == asset.id)).all()
    assert len(history) == 2
    assert sum(1 for item in history if item.is_open) == 1


def test_필수_열이_없으면_전체를_거부한다(db):
    content = _upload_file([["아무값"]], headers=["엉뚱한열"])
    result = import_assets(db, content)

    assert result.created == 0
    assert "필수 열이 없습니다" in result.errors[0]


def test_엑셀이_아닌_파일은_안내_메시지를_준다(db):
    result = import_assets(db, b"this is not an excel file")
    assert result.has_errors
    assert "엑셀 파일을 열 수 없습니다" in result.errors[0]


def test_빈_행은_조용히_건너뛴다(db):
    result = import_assets(db, _upload_file([_row(), [None] * len(HEADERS)]))
    assert result.created == 1 and result.skipped == 0


def test_자산_목록_다운로드(db, asset, employee):
    asset.holder_id = employee.id
    asset.status = "IN_USE"
    db.commit()

    wb = load_workbook(io.BytesIO(export_assets([asset])))
    ws = wb.active
    assert [cell.value for cell in ws[1]] == HEADERS

    row = {header: ws.cell(row=2, column=i + 1).value for i, header in enumerate(HEADERS)}
    assert row["자산번호"] == asset.asset_no
    assert row["상태"] == "사용중"
    assert row["사용자 사번"] == employee.emp_no


def test_다운로드한_파일을_그대로_다시_올릴_수_있다(db, asset):
    content = export_assets([asset])
    result = import_assets(db, content)

    assert result.updated == 1 and result.skipped == 0


def test_화면에서_엑셀_파일을_내려받는다(admin_client, asset):
    response = admin_client.get("/assets/export")
    assert response.status_code == 200
    assert "spreadsheetml" in response.headers["content-type"]
    assert response.content[:2] == b"PK"  # xlsx 는 zip 형식이다

    template = admin_client.get("/assets/template")
    assert template.status_code == 200
    assert template.content[:2] == b"PK"


def test_엑셀이_아닌_파일_업로드는_거부한다(admin_client):
    response = admin_client.post(
        "/assets/import", files={"file": ("data.csv", b"a,b,c", "text/csv")}
    )
    assert response.status_code == 400
    assert "엑셀 파일" in response.text
